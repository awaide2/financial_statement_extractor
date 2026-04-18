import os
import re
import math
import time
import numpy as np
import pandas as pd
import pytesseract
from pdf2image import convert_from_path
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from labels_v3 import *

DEBUG = True
REPORT_DIR = 'batch_reports'
DEFAULT_DPI_LIST = [320,300,280,260,220]
pd.options.display.float_format = '{:,.0f}'.format

INCOME_POSITIVE_KEYWORDS = ['revenue','sales','gross profit','profit before','net profit','net income','income tax expense','cost of revenue','cost of sales']
INCOME_NEGATIVE_KEYWORDS = ['total assets','total liabilities','total equity','equity and liabilities','statement of financial position','balance sheet','share capital','retained earnings','current assets','current liabilities']
ROW_REJECT_WORDS_GENERAL = ['basis of preparation','accounting policies','the accompanying notes','earnings per share']
ROW_REJECT_WORDS_REVENUE = ['deferred revenue','contract liabilities']
ROW_REJECT_WORDS_TAX = ['income tax payable','deferred tax','deferred tax net']
ROW_REJECT_WORDS_NET = ['other comprehensive','net of tax','cash flow hedge','remeasurement loss']

def ensure_dir(path):
    os.makedirs(path, exist_ok=True)

def safe_name(name):
    return re.sub(r'[^A-Za-z0-9._-]+','_',str(name))

def section_print(text, level=0):
    prefix = '    ' * level
    if level == 0:
        print('\n' + '=' * 80)
        print(text)
        print('=' * 80)
    else:
        print(f'{prefix}{text}')


def clean_number_str(val):
    s = str(val).strip()
    s = s.replace('—','-').replace('–','-').replace('−','-')
    s = s.replace('O','0') if re.fullmatch(r'[\(\)\-,.O0-9 ]+', s) else s
    s = s.replace(' ','')
    s = s.replace(',','')
    s = s.replace('*','')
    s = s.replace('SAR','')
    s = s.replace('SR','')
    s = s.strip()
    if s in ['','-','--','nan','None',None]:
        return 0.0
    if s.startswith('(') and s.endswith(')'):
        s = '-' + s[1:-1]
    s = s.replace('(','').replace(')','')
    try:
        return float(s)
    except Exception:
        return float('nan')

def fmt_num(x):
    if x is None:
        return 'nan'
    try:
        v = float(np.array(x).flatten()[0])
        if math.isnan(v):
            return 'nan'
        return f'{v:,.0f}'
    except Exception:
        return str(x)

def is_number_like(token):
    t = str(token).strip()
    if t == '':
        return False
    if re.search(r'\d', t) is None:
        return False
    return re.fullmatch(r'[\(\)\-–—−\d,\.\s]+', t) is not None

def is_note_like(token):
    t = str(token).strip()
    if re.fullmatch(r'\d{1,2}', t):
        return True
    if re.fullmatch(r'\(?\d{1,2}\)?', t):
        return True
    return False

def normalize_label(text):
    s = str(text).lower()
    s = s.replace('&','and')
    s = s.replace('—',' ').replace('–',' ').replace('-',' ')
    s = re.sub(r'[^a-z0-9 ]+',' ',s)
    s = re.sub(r'\s+',' ',s).strip()
    return s

def normalize_year(y):
    m = re.search(r'20\d{2}', str(y))
    return m.group(0) if m else str(y)

def extract_years_from_lines(lines):
    years = []
    for line in lines:
        for y in re.findall(r'\b20\d{2}\b', str(line)):
            if y not in years:
                years.append(y)
    if len(years) >= 2:
        return years[:2]
    current = time.localtime().tm_year
    return [str(current), str(current - 1)]

def convert_pdf_page_to_image(pdf_path, page_number=1, dpi=300):
    images = convert_from_path(pdf_path, dpi=dpi, first_page=page_number, last_page=page_number)
    return images[0] if images else None

def ocr_all_pdf_pages_tesseract(pdf_path, dpi=200, max_pages=20):
    images = convert_from_path(pdf_path, dpi=dpi)
    images = images[:max_pages] if max_pages else images
    pages_text = []
    for idx, img in enumerate(images, start=1):
        txt = pytesseract.image_to_string(img)
        pages_text.append((idx, txt))
    return pages_text

def score_income_page_text(text):
    t = normalize_label(text)
    score = 0
    pos_hits = []
    neg_hits = []
    for pat in INCOME_HEADER_PATTERNS:
        if re.search(pat, t, flags=re.I):
            score += 8
            pos_hits.append(f'header:{pat}')
    for kw in INCOME_POSITIVE_KEYWORDS:
        if kw in t:
            score += 2
            pos_hits.append(kw)
    for kw in INCOME_NEGATIVE_KEYWORDS:
        if kw in t:
            score -= 3
            neg_hits.append(kw)
    if 'other comprehensive income' in t:
        score += 1
        pos_hits.append('other comprehensive income')
    return score, pos_hits, neg_hits


def count_income_labels(text):
    text_l=text.lower()
    score=0
    for p in revenue_patterns:
        if re.search(p,text_l,re.IGNORECASE):
            score+=1
    for p in cost_of_sales_patterns:
        if re.search(p,text_l,re.IGNORECASE):
            score+=1
    for p in gross_profit_patterns:
        if re.search(p,text_l,re.IGNORECASE):
            score+=1
    for p in sga_patterns:
        if re.search(p,text_l,re.IGNORECASE):
            score+=1
    for p in operating_profit_patterns:
        if re.search(p,text_l,re.IGNORECASE):
            score+=1
    for p in finance_income_patterns:
        if re.search(p,text_l,re.IGNORECASE):
            score+=1
    for p in finance_cost_patterns:
        if re.search(p,text_l,re.IGNORECASE):
            score+=1
    for p in pretax_patterns:
        if re.search(p,text_l,re.IGNORECASE):
            score+=1
    for p in tax_patterns:
        if re.search(p,text_l,re.IGNORECASE):
            score+=1
    for p in net_income_patterns:
        if re.search(p,text_l,re.IGNORECASE):
            score+=1
    return score


# ============================================================
# 🧠 HYBRID ANCHOR FUSION (CONSENSUS + VALIDATION)
# ============================================================
def fuse_income_anchors(all_results):
    print("\n🔷 Running Anchor Fusion...")
    fused={}
    keys=['Revenue','CostOfSales','GrossProfitCalc','OperatingProfitCalc','FinanceCost','PreTaxCalc','TaxZakat','NetIncomeCalc']
    def is_valid(key,val,revenue_ref=None):
        if val is None:
            return False
        arr=np.array(val,dtype=float)
        if arr.size==0 or np.isnan(arr).all():
            return False
        if key=='Revenue':
            if np.nanmean(arr)<=0:
                return False
        if key=='CostOfSales':
            if np.nanmean(arr)>0:
                return False
        if revenue_ref is not None:
            rev=np.nanmean(np.abs(revenue_ref))
            val_mean=np.nanmean(np.abs(arr))
            if rev>0:
                if val_mean>rev*3:
                    return False
        return True
    for key in keys:
        candidates=[]
        for r in all_results:
            anchors=r.get("anchors") or {}
            val=anchors.get(key)
            if val is not None:
                candidates.append(val)
        if not candidates:
            print(f"⚠ Missing {key}")
            continue
        best=None
        best_count=0
        for c in candidates:
            count=sum(np.allclose(np.array(c),np.array(x),atol=1) for x in candidates)
            if count>best_count:
                best=c
                best_count=count
        revenue_ref=fused.get('Revenue')
        if not is_valid(key,best,revenue_ref):
            print(f"🚫 Rejected {key} (invalid)")
            continue
        fused[key]=best
        print(f"🟢 Fused {key}")
    return fused


def select_best_structure(engine_results):
    print("\n🧠 Selecting best structure across engines...")
    best_df=None
    best_source=None
    best_score=float("-inf")
    for engine_name,res in engine_results.items():
        if not res:
            continue
        df=res.get("df")
        if df is None or df.empty:
            continue
        try:
            # ============================================================
            # 🔵 TRY RAW VS CLEAN VERSION
            # ============================================================
            df_raw=df.copy()
            def _clean_df(df):
                try:
                    dfc=df.copy()
                    dfc=dfc.dropna(how='all')
                    dfc=dfc[dfc.iloc[:,0].astype(str).str.strip()!=""]
                    dfc=dfc[~dfc.iloc[:,0].astype(str).str.lower().str.contains("statement of|for the year ended|notes|company|director|officer",regex=True)]
                    return dfc.reset_index(drop=True)
                except:
                    return df
            df_clean=_clean_df(df_raw)
            candidates=[("RAW",df_raw),("CLEAN",df_clean)]
            for version,df2 in candidates:
                row_count=len(df2)
                if row_count==0:
                    continue
                item_col=df2.columns[0]
                value_cols=list(df2.columns[1:])
                if len(value_cols)==0:
                    continue
                numeric_full=0
                numeric_any=0
                bad_item_rows=0
                broken_value_rows=0
                for _,row in df2.iterrows():
                    item=str(row[item_col]).strip().lower()
                    vals=[str(row[c]).strip() for c in value_cols]
                    non_empty_vals=[v for v in vals if v not in ["","None","nan","NaN"]]
                    valid_num_count=0
                    broken_here=False
                    for v in non_empty_vals:
                        vv=v.replace("(","").replace(")","").replace(",","").replace("-","").replace(".","").strip()
                        if vv.isdigit():
                            valid_num_count+=1
                        else:
                            if any(ch.isdigit() for ch in v):
                                broken_here=True
                    if valid_num_count>=2:
                        numeric_full+=1
                    elif valid_num_count>=1:
                        numeric_any+=1
                    if broken_here:
                        broken_value_rows += 1

                    # ============================================================
                    # 🔴 FIX: SMART ROW FILTER (LABEL-DRIVEN)
                    # ============================================================
                    item_norm = normalize_label(item)
                    # ---- skip empty ----
                    if item_norm == "":
                        continue
                    # ---- skip non-alpha OCR junk ----
                    if not any(ch.isalpha() for ch in item_norm):
                        continue
                    # ---- skip known NON-INCOME content (USE LABEL SYSTEM) ----
                    if any(kw in item_norm for kw in INCOME_NEGATIVE_KEYWORDS):
                        continue
                    if any(bad in item_norm for bad in ROW_REJECT_WORDS_GENERAL):
                        continue
                    # ---- skip structural headers using patterns (CRITICAL) ----
                    header_patterns = [
                        r'statement of',
                        r'for the year ended',
                        r'notes to',
                        r'accounting policies' ]
                    if any(re.search(p, item_norm, re.IGNORECASE) for p in header_patterns):
                        continue
                    if item=="":
                        bad_item_rows+=1
                        continue
                    if any(x in item for x in ["statement of income","statement of profit","for the year ended","notes sar","notes","the accompanying notes","chief financial officer","board member","managing director","company","joint stock company"]):
                        bad_item_rows+=1
                        continue
                    if not any(ch.isalpha() for ch in item):
                        bad_item_rows+=1
                        continue
                        
                score=0
                score+=numeric_full*4.0
                score+=numeric_any*2.0
                score-=bad_item_rows*5.0
                score-=broken_value_rows*4.0
                if row_count>25:
                    score-=(row_count-25)*1.5
                score+=min(row_count,25)*0.2
                print(f"🔍 {engine_name} [{version}] → rows={row_count} | full={numeric_full} | any={numeric_any} | bad={bad_item_rows} | broken={broken_value_rows} | score={round(score,2)}")
                if score>best_score:
                    best_score=score
                    best_df=df2
                    best_source=f"{engine_name}_{version}"
        except Exception as e:
            print(f"⚠️ Structure scoring failed for {engine_name}: {e}")
    if best_df is not None:
        print(f"🟢 Selected Structure → {best_source}")
    else:
        print("⚠️ No structure selected")
    return best_df,best_source



def detect_income_statement_pages(pdf_path,max_pages=20,dpi=160):
    section_print('🧭 INCOME PAGE DETECTION')
    candidates=[]
    pages_text=ocr_all_pdf_pages_tesseract(pdf_path,dpi=dpi,max_pages=max_pages)
    for page_no,text in pages_text:

        text_l = text.lower()

        # ============================================================
        # 🔴 HARD BLOCKS (CRITICAL FIX)
        # ============================================================
        if re.search(r'cash flows?', text_l):
            print(f'--- Page {page_no}: ❌ skipped (cash flow detected)')
            continue
        if 'notes to the financial statements' in text_l:
            print(f'--- Page {page_no}: ❌ skipped (notes page)')
            continue
        if 'accounting policies' in text_l:
            print(f'--- Page {page_no}: ❌ skipped (policy page)')
            continue
        # ============================================================
        # 🔴 BLOCK PURE COMPREHENSIVE INCOME (STRONG FIX)
        # ============================================================
        if re.search(r'statement of comprehensive income', text_l):
            # allow combined statements only
            if not re.search(r'profit|loss', text_l):
                print(f'--- Page {page_no}: ❌ skipped (pure comprehensive income)')

                continue

        # ============================================================
        # 🔴 HEADER REQUIREMENT (NEW)
        # ============================================================
        has_header = any([
            re.search(r'statement of (profit|income)', text_l),
            re.search(r'profit or loss', text_l),
        ])

        # ============================================================
        # 🔴 NUMERIC DENSITY (NEW)
        # ============================================================
        numeric_count = len(re.findall(r'\d{1,3}(,\d{3})+', text))
        if not has_header and numeric_count < 8:
            print(f'--- Page {page_no}: ❌ skipped (weak header & low numbers)')
            continue
        # ============================================================
        # EXISTING SCORE (KEEP)
        # ============================================================
        score=count_income_labels(text)
        print(f'--- Page {page_no}: score={score} | nums={numeric_count} | header={has_header}')
        # ============================================================
        # 🔴 STRICT THRESHOLD (UPDATED)
        # ============================================================
        if score >= 2 or (has_header and numeric_count >= 8):
            candidates.append((page_no,score,numeric_count))
    # ============================================================
    # 🔴 SORT (IMPROVED)
    # ============================================================
    candidates=sorted(candidates,key=lambda x:(-x[1],-x[2],x[0]))
    chosen=[x[0] for x in candidates[:3]]
    if not chosen:
        chosen=[1]

    print(f'\n🧭 Selected income candidate page(s): {chosen}')

    return chosen



def extract_raw_lines(text):
    return [line.strip() for line in str(text).split('\n') if line.strip()]



def remove_note_column_if_present_from_tokens(parts):

    FULL_NUMBER_REGEX=r'\(?-?\d{1,3}(?:,\d{3})+(?:\.\d+)?\)?'

    cleaned_parts = []

    # ==== NOISY OCR NORMALIZATION (CONDITIONAL) ====
    joined_preview = ' '.join(parts)
    has_dot_thousands = bool(re.search(r'\d{1,3}(\.\d{3}){1,}', joined_preview))
    has_letter_noise = any(re.search(r'\d+[A-Za-z]+\d*', p) for p in parts)

    if has_dot_thousands or has_letter_noise:
        normalized_parts = []
        for p in parts:
            if re.fullmatch(r'\(?-?\d{1,3}(\.\d{3})+(\.\d+)?\)?', p):
                p = p.replace('.', ',')
            if re.search(r'\d', p):
                p = re.sub(r'(?<=\d)[A-Za-z](?=\d)', '', p)
                p = re.sub(r'[A-Za-z]', '', p)

                # ==== SAFE FIX: LARGE NUMBER FALSE DECIMAL ====
                m = re.fullmatch(r'\(?(-?\d{1,3}(?:,\d{3})+)\.(\d{3})\)?', p)
                if m:
                    int_part = m.group(1).replace(',', '')
                    if len(int_part) >= 7:
                        p = p.replace('.', ',')

            p = re.sub(r'\(\s*0+(\d)', r'(\1', p)
            normalized_parts.append(p)

        parts = normalized_parts

    # ============================================================
    # 🔵 STEP 1: BASIC CLEAN
    # ============================================================
    i=0
    while i<len(parts):
        cur=parts[i]
        if re.match(r'^\d{1,2},$',cur) and i+1<len(parts):
            cleaned_parts.append(cur+parts[i+1])
            i+=2
        else:
            cleaned_parts.append(cur)
            i+=1

    # ============================================================
    # 🔴 FIX: MERGE SPLIT NUMBERS (CORRECT PLACEMENT)
    # ============================================================
    merged_parts=[]
    i=0
    while i<len(cleaned_parts):
        cur=cleaned_parts[i]

        if i+1<len(cleaned_parts):
            nxt=cleaned_parts[i+1]

            if re.fullmatch(r'\d{1,3},\d{1,2}',cur) and re.fullmatch(r'\(?-?\d{1,3}(?:,\d{3})+\)?',nxt):
                print(f"🔵 Merging split number → {cur} + {nxt}")

                merged=cur.replace(',','')+nxt.replace(',','')
                try:
                    merged_val=f"{int(merged):,}"
                    merged_parts.append(merged_val)
                    i+=2
                    continue
                except:
                    pass

        merged_parts.append(cur)
        i+=1

    cleaned_parts=merged_parts

    # ============================================================
    # 🔥 RECONSTRUCT FULL NUMBERS
    # ============================================================
    joined_line=' '.join(cleaned_parts)
    full_numbers=re.findall(FULL_NUMBER_REGEX,joined_line)

    numeric_values=[]

    for p in cleaned_parts:

        if p in ['-','--']:
            numeric_values.append(p)
            continue

        cleaned_p=re.sub(r'[^\d,\.\-\(\)]','',p)
        cleaned_p=cleaned_p.replace(' ','')

        try:
            num_test=float(cleaned_p.replace(',','').replace('(','-').replace(')',''))
            if abs(num_test)<10 and len(re.findall(r'\d',cleaned_p))<=2 and is_note_like(cleaned_p):
                continue
        except:
            pass

        if re.search(r'\d{1,3}(,\d{3})+',cleaned_p) or re.fullmatch(r'\(?-?\d+\.\d+\)?',cleaned_p):
            numeric_values.append(cleaned_p)

    # ============================================================
    # 🔥 SAFE FULL NUMBER USAGE
    # ============================================================
    valid_full_numbers=[]
    for n in full_numbers:
        try:
            val=float(n.replace(',','').replace('(','-').replace(')',''))
            if abs(val)>1000:
                valid_full_numbers.append(n)
        except:
            continue

    if len(valid_full_numbers)>=2 and len(numeric_values)<2:
        print("🔵 Using reconstructed full numbers (safe)")
        numeric_values=valid_full_numbers

    # ============================================================
    # 🔥 FINAL CLEAN
    # ============================================================
    cleaned_numbers=[]
    for n in numeric_values:
        try:
            float(n.replace(',','').replace('(','-').replace(')',''))
            cleaned_numbers.append(n)
        except:
            continue
    numeric_values=cleaned_numbers
    # ============================================================
    # 🔥 REMOVE SMALL FIRST TOKEN
    # ============================================================
    if len(numeric_values)>=3:
        try:
            first_num=float(numeric_values[0].replace(',','').replace('(','-').replace(')',''))
            if abs(first_num)<100:
                numeric_values=numeric_values[1:]
        except:
            pass

    return numeric_values

def build_engine_result(engine, status, diff, major_hits, coverage, anchors, df, col_conf=0, pages_used=None, dpi=None):
    if pages_used is None:
        pages_used = []

    if df is None:
        df = pd.DataFrame()

    if anchors is None:
        anchors = {}

    return {
        'engine': engine,
        'status': status,
        'diff': float(diff),
        'major_hits': int(major_hits),
        'coverage': float(coverage),
        'anchors': anchors,
        'df': df,
        'col_conf': int(col_conf),
        'pages_used': pages_used,
        'dpi': dpi
    }


def parse_inline_lines_to_df(lines):
    lines=reconstruct_lines(lines)
    lines=force_vertical_pairing(lines)
    lines=reconstruct_blocks(lines)

    years=extract_years_from_lines(lines)
    formatted=[]

    for i,line in enumerate(lines):
        parts=line.split()
        numbers=remove_note_column_if_present_from_tokens(parts)

        # REMOVE ONLY DETECTED NUMBERS FROM LABEL (SAFE MODE)
        # AVOIDS BREAKING OCR TOKENS OR MIXED NUMBERS
        if len(numbers)>=1:
            label=line
            for num in numbers[:2]:
                label=label.replace(num,' ')

            label=re.sub(r'\b\d{1,2}\b',' ',label)
            label=re.sub(r'\s{2,}',' ',label).strip(' :-–—').strip()

            # ============================================================
            # 🔴 LABEL NORMALIZATION (CRITICAL FIX)
            # ============================================================
            label_norm = label.lower()


            # ---- Fix "gross profit" broken into 2 lines (STRONG FIX) ----
            if label_norm.strip() == "profit":
                prev_line = lines[i - 1].lower() if i > 0 else ""
                next_line = lines[i + 1].lower() if i + 1 < len(lines) else ""
                if "gross" in prev_line or "gross" in next_line:
                    print(f"🟢 Fixed label → Gross profit (strong)")

                    label = "Gross profit"

            # ---- Fix operating profit naming ----
            if "profit" in label_norm and "operations" in label_norm:
                label = "Operating profit"

            # ---- Fix revenue header corruption (SR *000 Revenue) ----
            if "revenue" in label_norm and ("sr" in label_norm or "*000" in label_norm):
                label = "Revenue"

            # ============================================================
            # 🔴 FINAL LABEL PROTECTION (SAFE)
            # ONLY keep explicit gross-profit repair from nearby context
            # ============================================================
            # ============================================================
            # ORIGINAL LOGIC (UNCHANGED BELOW)
            # ============================================================
            if re.search(r'[A-Za-z]', label):
                if len(numbers) >= 2:
                    formatted.append([label, numbers[0], numbers[1]])
                else:
                    # ============================================================
                    # 🔴 FIX: TRY RECOVER SECOND VALUE (INLINE + VERTICAL)
                    # ============================================================
                    raw_numbers = re.findall(r'\(?-?\d{1,3}(?:,\d{3})+(?:\.\d+)?\)?', str(line))
                    if len(raw_numbers) >= 2:
                        print(f"🟢 Recovered inline → {label}")
                        formatted.append([label, raw_numbers[0], raw_numbers[1]])
                    else:
                        # 🔴 NEW: CHECK NEXT LINES (VERTICAL STRUCTURE)
                        next_vals = []
                        for j in range(i + 1, min(i + 4, len(lines))):
                            nxt = lines[j]
                            nums = re.findall(r'\(?-?\d{1,3}(?:,\d{3})+(?:\.\d+)?\)?', str(nxt))
                            if nums:
                                next_vals.extend(nums)
                            if len(next_vals) >= 2:
                                break
                        if len(next_vals) >= 2:
                            print(f"🟢 Recovered vertical → {label}")
                            formatted.append([label, next_vals[0], next_vals[1]])
                        else:
                            formatted.append([label, numbers[0], None])


        elif re.search(r'[A-Za-z]',line):
            formatted.append([line,None,None])

    if not formatted:
        if len(years)<2:
            years=years+['Unknown']*(2-len(years))
        return pd.DataFrame(columns=['Item',years[0],years[1]])

    # FORCE CONSISTENT STRUCTURE
    # ENSURE ALL ROWS HAVE 3 COLUMNS
    normalized = []
    for row in formatted:
        if len(row) == 3:
            normalized.append(row)
        elif len(row) == 2:
            normalized.append([row[0], row[1], None])
    df = pd.DataFrame(normalized, columns=['Item', years[0], years[1]])
    # ============================================================
    # 🔴 REMOVE DUPLICATE / BROKEN LABELS (CRITICAL FIX)
    # ============================================================
    df['Item_clean'] = df['Item'].astype(str).str.lower().str.strip()
    # ---- drop standalone "profit" if "gross profit" exists ----
    if 'gross profit' in df['Item_clean'].values:
        before = len(df)
        df = df[df['Item_clean'] != 'profit']
        after = len(df)
        if before != after:
            print("🟢 Dropped duplicate 'profit' row → using 'Gross profit'")
    # ---- drop standalone "gross" row ----
    if 'gross profit' in df['Item_clean'].values:
        before = len(df)
        df = df[df['Item_clean'] != 'gross']
        after = len(df)
        if before != after:
            print("🟢 Dropped broken 'Gross' row")
    # ---- drop weak header artifacts that corrupt row search ----
    bad_exact_rows = [
        'sr 000',
        "sr '000",
        'sr *000',
        'note',
        'notes'
    ]
    df = df[~df['Item_clean'].isin(bad_exact_rows)]
    df = df.drop(columns=['Item_clean'])
    # ============================================================
    # 🔴 FINAL LABEL NORMALIZATION (CRITICAL FIX)
    # ============================================================
    df['Item'] = df['Item'].astype(str)
    df['Item'] = df['Item'].str.strip()
    # Fix broken "profit" → "Gross profit"
    df.loc[df['Item'].str.lower() == 'profit', 'Item'] = 'Gross profit'
    df = clean_statement_df(df)
    print(f"[DEBUG PARSE] rows={len(df)}")
    print(df.head(10))
    return df

def reconstruct_blocks(lines):
    """
    Group label + multiple numeric rows into structured rows
    Handles:
    - Label + note + multiple numbers
    - Shifted cost rows
    """
    result=[]
    i=0
    n=len(lines)

    while i<n:
        current=lines[i].strip()

        # detect label
        if not is_number_like(current):
            label=current
            numbers=[]
            j=i+1

            # collect next numeric lines
            while j<n and (is_number_like(lines[j]) or is_note_like(lines[j])):
                val=lines[j].strip()

                if is_number_like(val):
                    numbers.append(val)

                j+=1

            # ------------------------------------------------
            # 🔥 CASE 1: 2 numbers → normal row
            # ------------------------------------------------
            if len(numbers)>=2:
                result.append(f"{label} {numbers[0]} {numbers[1]}")

                # ------------------------------------------------
                # 🔥 CASE 2: EXTRA NUMBERS → SHIFTED NEXT ROW
                # ------------------------------------------------
                if len(numbers)>=4:
                    next_label=None
                    if j<n and not is_number_like(lines[j]):
                        next_label=lines[j].strip()
                        result.append(f"{next_label} {numbers[2]} {numbers[3]}")
                        j+=1

                i=j
                continue

        # default
        result.append(current)
        i+=1

    return result


def clean_statement_df(df):
    if df is None or df.empty:
        return df
    df = df.copy()
    df['Item'] = df['Item'].astype(str).str.replace(r'\s+\d+$','', regex=True)
    df['Item'] = df['Item'].astype(str).str.replace(r'\bnote\b',' ', case=False, regex=True)
    df['Item'] = df['Item'].astype(str).str.replace(r'[,\:\-]+$','', regex=True)
    df['Item'] = df['Item'].apply(lambda x: re.sub(r'\s+',' ', x).strip())
    df = df[df['Item'].str.contains(r'[A-Za-z]', na=False)]
    df = df[~df['Item'].str.fullmatch(r'20\d{2}', na=False)]
    df = df.drop_duplicates(subset=['Item'], keep='first').reset_index(drop=True)
    return df

def classify_extraction_format(lines):
    number_lines = [line for line in lines if len(re.findall(r'\d{1,3}(?:,\d{3})+', str(line))) >= 1]
    if len(number_lines) >= 4:
        return 'Inline Label & Numbers'
    return 'Unknown'

def row_is_rejected(item_text, label_keywords):
    s = normalize_label(item_text)
    for bad in ROW_REJECT_WORDS_GENERAL:
        if bad in s:
            return True
    label_set = {normalize_label(x) for x in label_keywords}
    if any(x in label_set for x in ['revenue','sales','total revenue','net sales','contract revenue','turnover']):
        if any(bad in s for bad in ROW_REJECT_WORDS_REVENUE):
            return True
    if any(x in label_set for x in ['zakat','zakat and income tax','income tax','tax expense','deferred tax','withholding tax']):
        if any(bad in s for bad in ROW_REJECT_WORDS_TAX):
            return True
    if any(x in label_set for x in ['net profit','net income','profit for the year','profit attributable','net earnings','net loss']):
        if any(bad in s for bad in ROW_REJECT_WORDS_NET):
            return True
    return False




# ============================================================
# 🔥 PATCHED get_row (SAFE — NO REGRESSION)
# Adds:
# - next-row fallback ONLY when current row is weak
# - stronger "net" prioritization
# ============================================================
def get_row(df, label_keywords, regex_patterns=None):

    if df is None or df.empty:
        return None

    df2 = df.copy()
    df2['Norm'] = df2['Item'].apply(normalize_label)
    df2 = df2[~df2['Item'].apply(lambda x: row_is_rejected(x, label_keywords))]

    # ---- helper ----
    def extract_values(row):
        year_cols = [c for c in df2.columns if c not in ['Item','Norm']]
        return np.array([clean_number_str(row[c]) for c in year_cols], dtype=float)

    def is_weak(values):
        if values is None or len(values) == 0:
            return True
        if np.isnan(values).all():
            return True
        return np.nansum(np.abs(values)) < 1e-5  # 🔥 critical threshold

    def try_next_row(row):
        idx = row.name
        if idx + 1 in df2.index:
            next_row = df2.loc[idx + 1]
            next_vals = extract_values(next_row)

            # only accept if clearly stronger
            if not is_weak(next_vals) and np.nansum(np.abs(next_vals)) > np.nansum(np.abs(extract_values(row))):
                print(f"🔁 Using next row fallback → {next_row['Item']}")
                return next_row
        return row

    def choose_best(candidate_df):
        if candidate_df is None or candidate_df.empty:
            return None

        # 🔥 PRIORITY 1: rows containing 'net'
        net_df = candidate_df[candidate_df['Norm'].str.contains(r'\bnet\b', na=False)]
        if not net_df.empty:
            candidate_df = net_df

        # 🔥 DIRECTLY SCORE ROWS (NO VALUE MATCHING)
        best_row = None
        best_score = -1

        for _, r in candidate_df.iterrows():
            if DEBUG:
                print(f"🔍 Row → {r['Item']}") 
                # ============================================================
                # 🔍 CRASH DEBUG (CRITICAL)
                # ============================================================
                #print(f"🔍 ROW INDEX → {r.name}")
                #print(f"🔍 ROW FULL → {r.to_dict()}")
                #print(f"🔍 DF COLUMNS → {list(df2.columns)}")
                year_cols = [c for c in df2.columns if c not in ['Item', 'Norm']]
                print(f"🔍 YEAR COLS → {year_cols}")

                try:
                    vals = [r[c] for c in year_cols]
                    print(f"🔍 SAFE VALUES → {vals}")
                except Exception as e:
                    print(f"🚨 VALUE ERROR → {e}")
                print(f"🔍 RAW VALUES → {[r[c] for c in year_cols]}")
                

            text = str(r['Item']).lower()
            # ============================================================
            # 🔴 STRICT NUMERIC COLUMN FIX (CRITICAL)
            # ============================================================
            year_cols = [c for c in df2.columns if c not in ['Item', 'Norm']]
            vals = np.array([clean_number_str(r[c]) for c in year_cols], dtype=float)
            if len(vals) == 0 or np.isnan(vals).all():
                continue
            magnitude = np.abs(vals).sum()
            # ============================================================
            # 🔴 IMPROVED SCORING (SIGN-AWARE FIX)
            # ============================================================
            score = magnitude
            # ---- PRIORITY BOOSTS ----
            if 'total' in text:
                score *= 3
            if 'profit' in text:
                score *= 2
            if 'net' in text:
                score *= 1.5
            # ============================================================
            # 🔴 SIGN VALIDATION (CRITICAL)
            # ============================================================
            if 'revenue' in text:
                if np.nanmean(vals) < 0:
                    print(f"🚫 Rejecting negative revenue → {text}")
                    score *= 0.1

            if 'cost' in text or 'expense' in text:
                if np.nanmean(vals) > 0:
                    print(f"🚫 Rejecting positive cost → {text}")
                    score *= 0.1

            # ---- PENALTY ----
            if 'medical' in text or 'pharmaceutical' in text:
                score *= 0.3

            if score > best_score:
                best_score = score
                best_row = r
        # 🔥 NEW: fallback ONLY if weak
        vals = extract_values(best_row)
        if is_weak(vals):
            best_row = try_next_row(best_row)

        return best_row

    # ============================================================
    # 1. EXACT MATCH
    # ============================================================
    matches = []
    for lbl in label_keywords:
        key = normalize_label(lbl)
        exact = df2[df2['Norm'] == key]
        if not exact.empty:
            matches.append(exact)

    if matches:
        candidate = pd.concat(matches).drop_duplicates()
        return choose_best(candidate)

    # ============================================================
    # 2. REGEX MATCH
    # ============================================================
    if regex_patterns:
        regex_hits = []
        for pat in regex_patterns:
            hit = df2[df2['Item'].str.contains(pat, case=False, na=False, regex=True)]
            if not hit.empty:
                regex_hits.append(hit)

        if regex_hits:
            candidate = pd.concat(regex_hits).drop_duplicates()
            return choose_best(candidate)

    # ============================================================
    # 3. LOOSE MATCH
    # ============================================================
    loose_hits = []
    for lbl in label_keywords:
        key = normalize_label(lbl)
        hit = df2[df2['Norm'].str.contains(key, na=False)]
        if not hit.empty:
            loose_hits.append(hit)

    if loose_hits:
        candidate = pd.concat(loose_hits).drop_duplicates()
        return choose_best(candidate)

    return None


# ==========================================
# SMART ROW SELECTION (TOTAL PRIORITY)
# ==========================================
def select_best_row(candidates, clean_func):
    candidates = candidates.copy()

    def score_row(row):
        text = str(row['Item']).lower()

        year_cols = [c for c in candidates.columns if c not in ['Item', 'Norm']]
        values = pd.Series([clean_func(row[c]) for c in year_cols])
        magnitude = values.abs().sum()

        score = magnitude

        # ---- PRIORITY BOOSTS ----
        if 'total' in text:
            score *= 3
        if 'profit' in text:
            score *= 2
        if 'net' in text:
            score *= 1.5

        # ---- PENALTY ----
        if 'medical' in text or 'pharmaceutical' in text:
            score *= 0.3  # avoid components

        return score
    candidates['score'] = candidates.apply(score_row, axis=1)
    best = candidates.sort_values('score', ascending=False).iloc[0]
    year_cols = [c for c in candidates.columns if c not in ['Item', 'Norm']]
    return [clean_func(best[c]) for c in year_cols]


def get_row_old(df, label_keywords, regex_patterns=None):
    if df is None or df.empty:
        return None
    df2 = df.copy()
    df2['Norm'] = df2['Item'].apply(normalize_label)
    df2 = df2[~df2['Item'].apply(lambda x: row_is_rejected(x, label_keywords))]
    matches = []
    for lbl in label_keywords:
        key = normalize_label(lbl)
        exact = df2[df2['Norm'] == key]
        if not exact.empty:
            matches.append(exact)
    if matches:
        candidate = pd.concat(matches).drop_duplicates()
        return choose_best_row(candidate)
    if regex_patterns:
        regex_hits = []
        for pat in regex_patterns:
            hit = df2[df2['Item'].str.contains(pat, case=False, na=False, regex=True)]
            if not hit.empty:
                regex_hits.append(hit)
        if regex_hits:
            candidate = pd.concat(regex_hits).drop_duplicates()
            return choose_best_row(candidate)
    loose_hits = []
    for lbl in label_keywords:
        key = normalize_label(lbl)
        hit = df2[df2['Norm'].str.contains(key, na=False)]
        if not hit.empty:
            loose_hits.append(hit)
    if loose_hits:
        candidate = pd.concat(loose_hits).drop_duplicates()
        return choose_best_row(candidate)
    return None

def reconstruct_lines(lines):
    """
        Rebuild OCR-broken rows:
        1) Merge vertical patterns: Label + number + number → single row
        2) Merge short multiline labels
        SAFE: Only triggers on clear patterns → won’t affect clean tables
        """
    if not lines:
        return lines
    reconstructed = []
    i = 0
    n = len(lines)
    while i < n:
        current = lines[i].strip()
        # ------------------------------------------------------------
        # 1️⃣ DETECT LABEL + STACKED NUMBERS (VERTICAL FIX)
        # Example:
        # Revenue
        # 63,322,248
        # 51,892,213
        # ------------------------------------------------------------
        if i + 2 < n:
            next1 = lines[i + 1].strip()
            next2 = lines[i + 2].strip()

            if (
                not is_number_like(current)
                and is_number_like(next1)
                and is_number_like(next2)
            ):
                merged = f"{current} {next1} {next2}"
                reconstructed.append(merged)
                i += 3
                continue
        # ------------------------------------------------------------
        # 2️⃣ MULTILINE LABEL MERGE (SAFE)
        # Only merge when:
        # - both lines are non-numeric
        # - short fragments (avoid merging full rows)
        # ------------------------------------------------------------
        if i + 1 < n:
            next_line = lines[i + 1].strip()

            if (
                    not is_number_like(current)
                    and not is_number_like(next_line)
                    and len(current.split()) <= 3
                    and len(next_line.split()) <= 3
                    and not re.search(r'income|expense|profit|loss', current, re.IGNORECASE)
                    and not re.search(r'income|expense|profit|loss', next_line, re.IGNORECASE)
            ):
                merged = f"{current} {next_line}"
                reconstructed.append(merged)
                i += 2
                continue
        # ------------------------------------------------------------
        # DEFAULT: KEEP LINE
        # ------------------------------------------------------------
        reconstructed.append(current)
        i += 1
    return reconstructed


def choose_best_row(candidate_df):
    if candidate_df is None or candidate_df.empty:
        return None
    year_cols = [c for c in candidate_df.columns if c not in ['Item','Norm']]
    scored = []
    for _, row in candidate_df.iterrows():
        vals = [clean_number_str(row[c]) for c in year_cols]
        abs_sum = sum(abs(v) for v in vals if not pd.isna(v))
        score = abs_sum
        norm = normalize_label(row['Item'])
        if 'profit for the period before' in norm:
            score += 5_000_000
        if 'profit for the period' in norm:
            score += 4_000_000
        if norm == 'gross profit':
            score += 3_000_000
        if norm == 'cost of revenue' or norm == 'cost of sales':
            score += 2_000_000
        if norm == 'revenue' or 'revenue from contracts' in norm:
            score += 2_000_000
        if 'net ' in norm or norm.startswith('net'):
            score += 1_000_000
        if 'total' in norm:
            score += 500_000
        scored.append((score, row))
    scored = sorted(scored, key=lambda x: x[0], reverse=True)
    return scored[0][1]

def row_to_values(row, year_cols):
    if row is None:
        return np.array([np.nan for _ in year_cols], dtype=float)
    vals = []
    for col in year_cols:
        vals.append(clean_number_str(row[col]))
    return np.array(vals, dtype=float)

def detect_sign_from_label(label):
    s = normalize_label(label)
    negative_terms = ['cost','expense','loss','zakat','tax','charges']
    positive_terms = ['revenue','income','profit','gain']
    if any(t in s for t in negative_terms):
        return -1
    if any(t in s for t in positive_terms):
        return 1
    return 1

def signed_row_values(row, year_cols):
    vals = row_to_values(row, year_cols)
    if row is None:
        return vals
    sign = detect_sign_from_label(row['Item'])
    return np.abs(vals) * sign

def compute_income_anchors_base(df):
    anchors = {}
    if df is None or df.empty:
        print("⚠ Empty DataFrame passed to compute_income_anchors")
        return anchors

    df2 = df.copy()
    # ============================================================
    # 🏦 BANK DETECTION (LABEL-DRIVEN)
    # ============================================================
    is_bank = False
    for lbl in bank_detection_labels:
        if df2['Item'].str.contains(lbl, case=False, na=False).any():
            is_bank = True
            break

    if not is_bank:
        for pat in bank_detection_patterns:
            if df2['Item'].str.contains(pat, case=False, na=False, regex=True).any():
                is_bank = True
                break

    # ============================================================
    # 🔥 NEW: MULTI-MODEL DETECTION (CRITICAL)
    # ============================================================
    model_type = "STANDARD"

    if is_bank:
        model_type = "BANK"
        print("🧠 MODEL DETECTED → BANK")

    anchors['ModelType'] = model_type

    # ============================================================
    # CLEAN LABELS
    # ============================================================
    df2['Item'] = df2['Item'].astype(str)
    df2['Item'] = df2['Item'].str.replace(r'[,\:\-\.\(\)]+$', '', regex=True)
    df2['Item'] = df2['Item'].str.replace(r'\s+', ' ', regex=True)
    df2['Item'] = df2['Item'].str.strip().str.lower()

    year_cols = list(df2.columns[1:])

    # ============================================================
    # 🔎 SHOW PREVIEW ONLY ONCE (AVOID SPAM)
    # ============================================================
    if DEBUG and not hasattr(compute_income_anchors_base, "_preview_printed"):
        section_print('🔎 Income Label Preview', 0)
        print(df2[['Item']].head(10).to_string(index=False))
        compute_income_anchors_base._preview_printed = True

    def debug_get(name, labels, patterns):
        row = get_row(df2, labels, patterns)
        if row is not None and not row.empty:
            print(f"  ✔ {name:<12} → {row['Item']}")
        else:
            print(f"  • {name:<12} → not found")
        return row

    # =========================
    # STANDARD ANCHORS (MODEL-AWARE FIX)
    # =========================
    if is_bank:
        rev_row = debug_get('Revenue (Bank)', bank_revenue_labels, bank_revenue_patterns)
    else:
        rev_row = debug_get('Revenue', revenue_labels, revenue_patterns)
    gross_row = debug_get('Gross Profit', gross_profit_labels, gross_profit_patterns)
    # ============================================================
    # SIMPLE COST SELECTION (STABLE)
    # ============================================================

    cos_row = get_row(df2, cost_of_sales_labels, cost_of_sales_patterns)

    # fallback: cost of operations (ONLY if nothing found)
    if cos_row is None:
        alt = df2[df2['Item'].str.contains(r'cost.*operations', case=False, na=False)]
        if not alt.empty:
            cos_row = alt.iloc[0]

    # ============================================================
    # 🔥 SG&A (AGGREGATED — SAFE FIX)
    # ============================================================
    sga_values = np.zeros(len(year_cols))

    for _, row in df2.iterrows():
        item = row['Item']

        if any(re.search(p, item, re.IGNORECASE) for p in sga_patterns):

            if any(x in item for x in ['research', 'depreciation', 'amortization', 'impairment']):
                continue

            if 'total' in item:
                continue

            vals = signed_row_values(row, year_cols)

            if not np.isnan(vals).all():
                sga_values += np.nan_to_num(vals)

    if np.all(sga_values == 0):
        anchors['SGA'] = np.array([np.nan]*len(year_cols))
        print("  • SGA          → not found")
    else:
        anchors['SGA'] = sga_values
        print(f"  ✔ SGA          → aggregated ({fmt_num(sga_values)})")

    # ============================================================
    # 🔥 NEW: EXTRA OPERATING COMPONENTS (SAFE ADD)
    # ============================================================
    rd_values = np.zeros(len(year_cols))
    dep_values = np.zeros(len(year_cols))
    imp_values = np.zeros(len(year_cols))

    for _, row in df2.iterrows():
        item = row['Item']
        vals = signed_row_values(row, year_cols)

        if np.isnan(vals).all():
            continue

        if re.search(r'research.*development', item, re.IGNORECASE):
            rd_values += np.nan_to_num(vals)

        if re.search(r'depreciation|amortization', item, re.IGNORECASE):
            dep_values += np.nan_to_num(vals)

        if re.search(r'impairment', item, re.IGNORECASE):
            imp_values += np.nan_to_num(vals)

    anchors['R&D'] = rd_values if not np.all(rd_values == 0) else np.array([np.nan]*len(year_cols))
    anchors['Depreciation'] = dep_values if not np.all(dep_values == 0) else np.array([np.nan]*len(year_cols))
    anchors['Impairment'] = imp_values if not np.all(imp_values == 0) else np.array([np.nan]*len(year_cols))

    # =========================
    # CONTINUE NORMAL FLOW
    # =========================
    op_inc_other_row = debug_get('Other Income', operating_other_income_labels, operating_other_income_patterns)
    op_exp_other_row = debug_get('Other Expense', operating_other_expense_labels, operating_other_expense_patterns)
    op_row = debug_get('Operating Profit', operating_profit_labels, operating_profit_patterns)
    fin_income_row = debug_get('Finance Income', finance_income_labels, finance_income_patterns)
    fin_cost_row = debug_get('Finance Cost', finance_cost_labels, finance_cost_patterns)
    assoc_row = debug_get('Associate', associate_labels, associate_patterns)
    pretax_row = debug_get('PreTax', pretax_labels, pretax_patterns)
    tax_row = debug_get('Tax', tax_labels, tax_patterns)
    # ============================================================
    # 🔴 FIX: BLOCK "BEFORE ZAKAT" MISCLASSIFICATION
    # ============================================================
    if tax_row is not None:
        label_text = str(tax_row['Item']).lower()

        # ------------------------------------------------------------
        # 🔴 BLOCK PRETAX MISCLASSIFICATION
        # ------------------------------------------------------------
        if 'before zakat' in label_text or 'before tax' in label_text:
            print(f"🚫 Ignoring invalid tax row → {tax_row['Item']}")
            tax_row = None

        # ------------------------------------------------------------
        # 🔴 EXTRA SAFETY: TAX SHOULD BE SMALLER THAN PRETAX
        # ------------------------------------------------------------
        if tax_row is not None and pretax_row is not None:
            year_cols = list(df2.columns[1:])  # ✅ FIX (define locally)
            pretax_vals = row_to_values(pretax_row, year_cols)
            tax_vals = row_to_values(tax_row, year_cols)

            if np.nansum(np.abs(tax_vals)) >= np.nansum(np.abs(pretax_vals)):
                print(f"🚫 Tax too large → likely misclassified ({tax_row['Item']})")
                tax_row = None
    net_row = debug_get('Net Income', net_income_labels, net_income_patterns)
    # ============================================================
    # ✅ SAFE REVENUE SELECTION (NO AGGREGATION)
    # ============================================================
    anchors['Revenue'] = row_to_values(rev_row, year_cols)
    print(f"🟢 Revenue selected → {fmt_num(anchors['Revenue'])}")
    # ============================================================
    # ✅ SAFE COST SELECTION (NO AGGREGATION)
    # ============================================================
    anchors['CostOfSales'] = signed_row_values(cos_row, year_cols)
    print(f"🟢 Cost selected → {fmt_num(anchors['CostOfSales'])}")


    anchors['GrossProfitReported'] = row_to_values(gross_row, year_cols)
    anchors['GrossProfitCalc'] = np.nan_to_num(anchors['Revenue']) + np.nan_to_num(anchors['CostOfSales'])

    anchors['OperatingOtherIncome'] = signed_row_values(op_inc_other_row, year_cols)
    anchors['OperatingOtherExpense'] = signed_row_values(op_exp_other_row, year_cols)

    anchors['OperatingProfitReported'] = row_to_values(op_row, year_cols)

    # ============================================================
    # 🔥 UPDATED OPERATING CALC
    # ============================================================
    if not np.isnan(anchors['OperatingProfitReported']).all():
        anchors['OperatingProfitCalc'] = anchors['OperatingProfitReported']
    else:
        anchors['OperatingProfitCalc'] = (
            np.nan_to_num(anchors['GrossProfitReported']) +
            np.nan_to_num(anchors['SGA']) +
            np.nan_to_num(anchors['R&D']) +
            np.nan_to_num(anchors['Depreciation']) +
            np.nan_to_num(anchors['Impairment']) +
            np.nan_to_num(anchors['OperatingOtherIncome']) +
            np.nan_to_num(anchors['OperatingOtherExpense'])
        )

    anchors['FinanceIncome'] = signed_row_values(fin_income_row, year_cols)
    anchors['FinanceCost'] = signed_row_values(fin_cost_row, year_cols)
    anchors['Associate'] = signed_row_values(assoc_row, year_cols)

    anchors['PreTaxReported'] = row_to_values(pretax_row, year_cols)

    if not np.isnan(anchors['PreTaxReported']).all():
        anchors['PreTaxCalc'] = anchors['PreTaxReported']
    else:
        anchors['PreTaxCalc'] = (
            np.nan_to_num(anchors['OperatingProfitReported']) +
            np.nan_to_num(anchors['FinanceIncome']) +
            np.nan_to_num(anchors['FinanceCost']) +
            np.nan_to_num(anchors['Associate'])
        )

    anchors['TaxZakat'] = signed_row_values(tax_row, year_cols)
    anchors['NetIncomeReported'] = row_to_values(net_row, year_cols)

    if not np.isnan(anchors['NetIncomeReported']).all():
        anchors['NetIncomeCalc'] = anchors['NetIncomeReported']
    else:
        anchors['NetIncomeCalc'] = (
            np.nan_to_num(anchors['PreTaxCalc']) +
            np.nan_to_num(anchors['TaxZakat'])
        )

    if DEBUG:
        section_print('📊 INCOME DEBUG SNAPSHOT', 0)
        print(f"Revenue              : {fmt_num(anchors['Revenue'])}")
        print(f"R&D                  : {fmt_num(anchors.get('R&D'))}")
        print(f"Depreciation         : {fmt_num(anchors.get('Depreciation'))}")
        print(f"Impairment           : {fmt_num(anchors.get('Impairment'))}")
        print(f"CostOfSales          : {fmt_num(anchors['CostOfSales'])}")
        print(f"GrossProfitReported  : {fmt_num(anchors['GrossProfitReported'])}")
        print(f"GrossProfitCalc      : {fmt_num(anchors['GrossProfitCalc'])}")
        print(f"SGA                  : {fmt_num(anchors['SGA'])}")
        print(f"FinanceCost          : {fmt_num(anchors['FinanceCost'])}")
        print(f"PreTaxReported       : {fmt_num(anchors['PreTaxReported'])}")
        print(f"PreTaxCalc           : {fmt_num(anchors['PreTaxCalc'])}")
        print(f"TaxZakat             : {fmt_num(anchors['TaxZakat'])}")
        print(f"NetIncomeReported    : {fmt_num(anchors['NetIncomeReported'])}")
        print(f"NetIncomeCalc        : {fmt_num(anchors['NetIncomeCalc'])}")

    return anchors



# ============================================================
# WRAPPER (SAFE COLUMN FIX)
# ============================================================
def compute_income_anchors(df):
    if df is None or df.empty:
        return {}
    # ---- ORIGINAL ----
    anchors_orig = compute_income_anchors_base(df)
    diff_orig = compute_income_diff(anchors_orig)
    # ---- SWAP ----
    if df.shape[1] >= 3:
        df_swapped = df.copy()
        cols = list(df.columns)
        df_swapped[cols[1]], df_swapped[cols[2]] = df[cols[2]], df[cols[1]]

        anchors_swap = compute_income_anchors_base(df_swapped)
        diff_swap = compute_income_diff(anchors_swap)

        print(f"🔍 Column Swap Check | Orig={diff_orig} | Swap={diff_swap}")

        if diff_swap < diff_orig:
            print("✅ Using SWAPPED columns")
            return anchors_swap

    # ---- DEFAULT ----
    return anchors_orig





def validate_buffett_inputs(anchors):
    if DEBUG:
        print("\n==============================")
        print("🧠 Buffett Input Validation")
        print("==============================")
    required=['Revenue','OperatingProfitCalc','NetIncomeCalc','FinanceCost']
    important=['GrossProfitReported','SGA','PreTaxCalc']
    result={'status':None,'missing_required':[],'missing_important':[]}
    # ============================================================
    # CHECK REQUIRED
    # ============================================================
    for k in required:
        v=anchors.get(k)
        if v is None or (isinstance(v, (list, np.ndarray)) and (np.isnan(v).all() or np.nansum(np.abs(v)) == 0 )):
            print(f"🔴 Missing (Critical) → {k}")
            result['missing_required'].append(k)
        else:
            print(f"🟢 OK → {k}")
    # ============================================================
    # CHECK IMPORTANT
    # ============================================================
    for k in important:
        v=anchors.get(k)
        if v is None or (isinstance(v,(list,np.ndarray)) and np.isnan(v).all()):
            print(f"🟡 Missing (Important) → {k}")
            result['missing_important'].append(k)
        else:
            print(f"🟢 OK → {k}")
    # ============================================================
    # FINAL STATUS
    # ============================================================
    if len(result['missing_required'])==0:
        if len(result['missing_important'])==0:
            print("\n✅ Buffett Analysis Ready")
            result['status']='READY'
        else:
            print("\n⚠️ Buffett Partial (some metrics limited)")
            result['status']='PARTIAL'
    else:
        print("\n❌ Buffett Analysis NOT POSSIBLE")
        result['status']='FAILED'
    return result



def validate_saii_inputs(anchors):
    if DEBUG:
        print("\n==============================")
        print("🧠 SAII Input Validation")
        print("==============================")
    required=['Revenue','OperatingProfitCalc','NetIncomeCalc','PreTaxCalc','TaxZakat','FinanceCost']
    result={'status':None,'missing_required':[]}
    # ============================================================
    # CHECK REQUIRED
    # ============================================================
    for k in required:
        v=anchors.get(k)
        if v is None or (isinstance(v,(list,np.ndarray)) and np.isnan(v).all()):
            print(f"🔴 Missing → {k}")
            result['missing_required'].append(k)
        else:
            print(f"🟢 OK → {k}")
    # ============================================================
    # FINAL STATUS
    # ============================================================
    if len(result['missing_required'])==0:
        print("\n✅ SAII Ready")
        result['status']='READY'
    elif len(result['missing_required'])<=2:
        print("\n⚠️ SAII Partial")
        result['status']='PARTIAL'
    else:
        print("\n❌ SAII NOT POSSIBLE")
        result['status'] = 'FAILED'
    return result




def _pair_diff(a, b):
    if a is None or b is None:
        return math.inf
    a = np.array(a, dtype=float)
    b = np.array(b, dtype=float)
    if len(a) == 0 or len(b) == 0:
        return math.inf
    if np.isnan(a).all() or np.isnan(b).all():
        return math.inf
    mask = ~(np.isnan(a) | np.isnan(b))
    if not np.any(mask):
        return math.inf
    return float(np.sum(np.abs(a[mask] - b[mask])))


# ============================================================
# 🔥 FAKE ZERO ANCHOR DETECTION (SAFE)
# ============================================================
def has_fake_zero_anchors(anchors):
    critical = ['Revenue','CostOfSales','GrossProfitReported','NetIncomeReported']
    valid_count = 0
    zero_count = 0
    for k in critical:
        v = anchors.get(k)
        if v is None:
            continue
        arr = np.array(v, dtype=float)
        if np.isnan(arr).all():
            continue
        valid_count += 1
        if np.nansum(np.abs(arr)) == 0:
            zero_count += 1
    if valid_count >= 2 and zero_count == valid_count:
        print("🚨 Fake zero anchors detected")
        return True
    return False


def compute_income_diff(anchors):

    # ============================================================
    # 🔥 FAKE ZERO DETECTION (CRITICAL FIX)
    # ============================================================
    if has_fake_zero_anchors(anchors):
        print("🚨 Forcing diff = INF due to fake anchors")
        return float('inf')

    pairs = [('GrossProfitCalc','GrossProfitReported'),
             ('OperatingProfitCalc','OperatingProfitReported'),
             ('PreTaxCalc','PreTaxReported'),
             ('NetIncomeCalc','NetIncomeReported')]
    diffs = []
    for left, right in pairs:
        left_val = anchors.get(left)
        right_val = anchors.get(right)
        # 🚨 DO NOT FORCE MATCH — INVALID PAIR
        if right_val is None or (isinstance(right_val, (list, np.ndarray)) and np.isnan(right_val).all()):
            continue
        diff = _pair_diff(left_val, right_val)
        if math.isfinite(diff):
            diffs.append(diff)
    if not diffs:
        return float('inf')
    return float(sum(diffs))


def major_hits_from_anchors(anchors,tol=2.0):
    hits=0
    for left,right in [('GrossProfitCalc','GrossProfitReported'),('OperatingProfitCalc','OperatingProfitReported'),('PreTaxCalc','PreTaxReported'),('NetIncomeCalc','NetIncomeReported')]:
        left_val=anchors.get(left)
        right_val=anchors.get(right)
        if right_val is None or (isinstance(right_val,(list,np.ndarray)) and np.isnan(right_val).all()):
            right_val=left_val
        # 🔴 NEW: VALIDITY CHECK (CRITICAL)
        def is_valid(x):
            if x is None:
                return False
            if isinstance(x,(list,np.ndarray)):
                return any(pd.notna(v) for v in x)
            return pd.notna(x)
        if not is_valid(left_val) and not is_valid(right_val):
            continue
        diff=_pair_diff(left_val,right_val)
        if math.isfinite(diff) and diff<=tol:
            hits+=1
    return hits




def coverage_from_df(df):
    found = 0
    missing = []
    if df is None or df.empty:
        return {'matched': 0, 'total': len(required_income_labels), 'coverage_percent': 0.0, 'missing_labels': list(required_income_labels)}
    norm_series = df['Item'].astype(str).apply(normalize_label)
    for lbl in required_income_labels:
        if norm_series.str.contains(normalize_label(lbl), case=False, na=False).any():
            found += 1
        else:
            missing.append(lbl)
    total = len(required_income_labels)
    return {'matched': found, 'total': total, 'coverage_percent': round(found * 100 / total, 2), 'missing_labels': missing}


# ============================================================
# 🧠 ANCHOR CONFIDENCE SCORING
# ============================================================
def compute_anchor_confidence(anchors):
    score=0
    rev=anchors.get('Revenue',[None,None])
    cost=anchors.get('CostOfSales',[None,None])
    gross=anchors.get('GrossProfitCalc',[None,None])
    def to_float(x):
        try:
            return float(str(x).replace(',',''))
        except:
            return None
    for r,c,g in zip(rev,cost,gross):
        r_val=to_float(r)
        c_val=to_float(c)
        g_val=to_float(g)
        if r_val is None or c_val is None:
            continue
        # ===============================
        # 🟢 Revenue dominance check
        # ===============================
        if abs(r_val)>abs(c_val):
            score+=1
        # ===============================
        # 🟢 Gross consistency check
        # ===============================
        if g_val is not None:
            if abs((r_val+c_val)-g_val)<abs(r_val)*0.2:
                score+=1
        # ===============================
        # 🟢 Positive revenue preference
        # ===============================
        if r_val>0:
            score+=1
    return score

def classify_status(diff, major_hits, col_conf, coverage_percent, row_count):

    print(f"[STATUS DEBUG] rows={row_count} | coverage={coverage_percent}% | hits={major_hits} | diff={diff} | col_conf={col_conf}")

    # ============================================================
    # 🔴 HARD FAIL — ONLY REAL STRUCTURE FAILS
    # ============================================================
    if row_count < 5 or col_conf == 0:
        print(f"[STATUS DEBUG] ❌ HARD FAIL → rows={row_count}, coverage={coverage_percent}%, col_conf={col_conf}")
        return 'STRUCTURE_BAD'

    # ============================================================
    # 🔴 FAKE DATA PROTECTION (KEEP)
    # ============================================================
    if major_hits >= 3 and diff == 0 and coverage_percent < 20:
        print("[STATUS DEBUG] ❌ FAKE PASS BLOCKED → strong math but weak structure")
        return 'STRUCTURE_BAD'

    # ============================================================
    # 🟢 SUCCESS_CONSOLIDATED (HIGH CONFIDENCE)
    # ============================================================
    if major_hits >= 3 and diff <= 1000 and col_conf >= 1:
        if coverage_percent >= 60:
            print("[STATUS DEBUG] ✅ SUCCESS_CONSOLIDATED (strong coverage)")
            return 'SUCCESS_CONSOLIDATED'
        else:
            print("[STATUS DEBUG] 🟡 SUCCESS_STRUCTURE_ONLY (low coverage but strong math)")
            return 'SUCCESS_STRUCTURE_ONLY'

    # ============================================================
    # 🟡 SUCCESS_STRUCTURE_ONLY (PARTIAL)
    # ============================================================
    if major_hits >= 2 and col_conf >= 1:
        print("[STATUS DEBUG] 🟡 SUCCESS_STRUCTURE_ONLY")
        return 'SUCCESS_STRUCTURE_ONLY'

    # ============================================================
    # 🔴 DEFAULT
    # ============================================================
    print("[STATUS DEBUG] ❌ DEFAULT → STRUCTURE_BAD")
    return 'STRUCTURE_BAD'



# ============================================================
# 🧠 ENGINE PENALTY SCORE (FLAT — NO NESTING)
# ============================================================

def compute_engine_penalty(result):

    print("🔵 Applying penalties...")

    penalty = 0

    anchors = result.get("anchors", {})

    revenue = anchors.get("Revenue")
    cost = anchors.get("CostOfSales")
    gross_rep = anchors.get("GrossProfitReported")
    gross_calc = anchors.get("GrossProfitCalc")

    # ------------------------------------------------------------
    # 🚫 Revenue invalid
    # ------------------------------------------------------------
    if revenue is None:
        print("🟤 Penalty → Revenue missing")
        penalty += 10
    else:
        arr = np.array(revenue, dtype=float)
        if np.isnan(arr).all() or np.nansum(np.abs(arr)) == 0:
            print("🟤 Penalty → Revenue zero/invalid")
            penalty += 10

    # ------------------------------------------------------------
    # 🚫 Cost wrong sign
    # ------------------------------------------------------------
    if cost is not None:
        arr = np.array(cost, dtype=float)
        if not np.isnan(arr).all():
            if np.nanmean(arr) > 0:
                print("🟤 Penalty → Cost should be negative")
                penalty += 10

    # ------------------------------------------------------------
    # 🚫 Gross mismatch
    # ------------------------------------------------------------
    if gross_rep is not None and gross_calc is not None:
        diff_val = _pair_diff(gross_rep, gross_calc)
        if math.isfinite(diff_val) and diff_val > 1000:
            print("🟤 Penalty → Gross mismatch")
            penalty += 8

    # ------------------------------------------------------------
    # 🚫 Too many NaNs
    # ------------------------------------------------------------
    nan_count = 0
    total = 0

    for v in anchors.values():
        if isinstance(v, (list, np.ndarray)):
            total += len(v)
            nan_count += np.isnan(v).sum()
    if total > 0:
        if nan_count / total > 0.5:
            print("🟤 Penalty → Too many NaNs")
            penalty += 5

    print(f"🟢 Total Penalty = {penalty}")

    return penalty



def result_score(result):

    status_rank = {
        'SUCCESS_CONSOLIDATED': 3,
        'SUCCESS_STRUCTURE_ONLY': 2,
        'STRUCTURE_BAD': 1
    }

    row_count = len(result.get('df')) if result.get('df') is not None else 0

    diff = result.get('diff')
    diff_score = -1e12 if diff is None or not math.isfinite(diff) else -float(diff)

    # ============================================================
    # 🧠 APPLY PENALTY (FLAT CALL)
    # ============================================================
    penalty = compute_engine_penalty(result)
    # base score from hits
    hits = int(result.get('major_hits', 0))
    final_score = hits * 2 - penalty
    return (
        status_rank.get(result.get('status'), 0),
        final_score,
        hits,
        int(result.get('col_conf', 0)),
        diff_score,
        row_count
    )

def force_vertical_pairing(lines):
    # Merge vertical OCR structure: label + number + number into one line
    # Fixes cases where numbers are below labels instead of inline
    result=[]
    i=0
    n=len(lines)
    while i<n:
        current=lines[i].strip()
        if not is_number_like(current):
            nums=[]
            j=i+1
            while j<n and is_number_like(lines[j]):
                nums.append(lines[j].strip())
                j+=1
            if len(nums)>=2:
                result.append(f"{current} {nums[0]} {nums[1]}")
                i=j
                continue
        result.append(current)
        i+=1
    return result


def print_anchor_debug(anchors):
    section_print('📌 ANCHOR SNAPSHOT', 0)
    for k, v in anchors.items():
        try:
            print(f'{k:<24}: {np.array(v).tolist()}')
        except Exception:
            print(f'{k:<24}: {v}')

def save_debug_case(pdf_path, pdf_result, easy_result, tess_result, inline_result, final_result, runtime_debug):
    ensure_dir(REPORT_DIR)
    out_path = os.path.join(REPORT_DIR, safe_name(os.path.basename(pdf_path).replace('.pdf','')) + '_income_debug.txt')
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('\n' + '=' * 80 + '\n')
        f.write('HYBRID DEBUG SNAPSHOT\n')
        f.write('=' * 80 + '\n\n')
        f.write(f'PDF: {pdf_path}\n')
        f.write(f'FINAL ENGINE: {final_result.get("engine")}\n')
        f.write(f'FINAL STATUS: {final_result.get("status")}\n')
        f.write(f'FINAL DIFF: {final_result.get("diff")}\n')
        f.write(f'DPI: {final_result.get("dpi")}\n')
        f.write(f'RUNTIME SEC: {final_result.get("runtime_sec")}\n')
        for title, obj in [('PDF_TEXT', pdf_result), ('EASYOCR', easy_result), ('TESSERACT', tess_result), ('INLINE', inline_result), ('FINAL_SELECTED', final_result)]:
            f.write('\n\n' + '=' * 90 + '\n')
            f.write(title + '\n')
            f.write('=' * 90 + '\n')

            if not obj:
                f.write('None\n')
                continue

            for key in ['engine', 'status', 'diff', 'major_hits', 'col_conf', 'dpi', 'pages_used', 'runtime_sec']:
                f.write(f'{key}: {obj.get(key)}\n')

            df = obj.get('df')
            if df is not None and not df.empty:
                f.write('\nDATAFRAME\n')
                f.write(df.to_string(index=False))
                f.write('\n')

            anchors = obj.get('anchors', {})
            if anchors:
                f.write('\nANCHORS\n')
                for k, v in anchors.items():
                    try:
                        f.write(f'{k}: {np.array(v).tolist()}\n')
                    except Exception:
                        f.write(f'{k}: {v}\n')
        f.write('\n\n' + '=' * 80 + '\n')
        f.write('FULL RUNTIME LOG\n')
        f.write('=' * 80 + '\n\n')
        f.write(runtime_debug)
    return out_path

def _safe_float_for_excel(v):
    try:
        x = float(v)
        if math.isnan(x) or math.isinf(x):
            return None
        return x
    except Exception:
        return v

def build_income_excel(result, pdf_path):
    ensure_dir(REPORT_DIR)
    base = safe_name(os.path.basename(pdf_path).replace('.pdf',''))
    out_path = os.path.join(REPORT_DIR, base + '_income_audit.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Extracted'
    meta = wb.create_sheet('Meta')
    # anchor_ws = wb.create_sheet('Anchors')

    # ============================================================
    # 📊 NEW: AUDIT SHEET (INCOME QUALITY)
    # ============================================================
    audit_ws = wb.create_sheet('Audit')
    df = result.get('df')
    if df is not None and not df.empty:
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))
        for col_idx in range(1, len(df.columns) + 1):
            ws.cell(1, col_idx).font = Font(bold=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = 28 if col_idx == 1 else 18
    meta_rows = [('PDF', os.path.basename(pdf_path)),('SelectedEngine', result.get('engine')),('Status', result.get('status')),('Diff', result.get('diff')),('MajorHits', result.get('major_hits')),('ColConf', result.get('col_conf')),('DPI', result.get('dpi')),('PagesUsed', ', '.join(map(str, result.get('pages_used', [])))),('RuntimeSec', result.get('runtime_sec'))]
    for r in meta_rows:
        meta.append(r)
    meta.column_dimensions['A'].width = 24
    meta.column_dimensions['B'].width = 60
    coverage = coverage_from_df(df) if df is not None else {'coverage_percent':0,'missing_labels':[]}
    meta.append(('CoveragePercent', coverage['coverage_percent']))
    meta.append(('MissingLabels', ', '.join(coverage['missing_labels'])))
    anchors = result.get('anchors', {})


    # ============================================================
    # 📊 AUDIT SHEET BUILD (CRITICAL)
    # ============================================================

    year_cols = [c for c in df.columns if c != 'Item'] if df is not None else []
    audit_ws.append(['Section', 'Metric'] + year_cols + ['Check', 'Diff'])

    def safe_values(a, year_count):
        if isinstance(a, (list, tuple, np.ndarray)):
            vals = list(np.array(a).tolist())
            while len(vals) < year_count:
                vals.append(None)
            return [_safe_float_for_excel(v) for v in vals[:year_count]]
        return [None] * year_count

    def diff_pair(a, b):
        try:
            a = np.array(a, dtype=float)
            b = np.array(b, dtype=float)
            return np.nansum(np.abs(a - b))
        except:
            return None

    def check_flag(diff):
        if diff is None:
            return 'N/A'
        return '✔' if diff <= 1 else '✖'

    pairs = [
        ('Revenue', anchors.get('Revenue')),
        ('Cost', anchors.get('CostOfSales')),
        ('Gross Reported', anchors.get('GrossProfitReported')),
        ('Gross Calc', anchors.get('GrossProfitCalc')),
        ('Operating', anchors.get('OperatingProfitReported')),
        ('PreTax Reported', anchors.get('PreTaxReported')),
        ('PreTax Calc', anchors.get('PreTaxCalc')),
        ('Tax', anchors.get('TaxZakat')),
        ('Net Reported', anchors.get('NetIncomeReported')),
        ('Net Calc', anchors.get('NetIncomeCalc'))
    ]

    year_count = len(year_cols)
    for name, val in pairs:
        vals = safe_values(val, year_count)
        audit_ws.append(['Core', name] + vals + ['', ''])

    validations = [
        ('Gross Check', 'GrossProfitReported', 'GrossProfitCalc'),
        ('PreTax Check', 'PreTaxReported', 'PreTaxCalc'),
        ('Net Check', 'NetIncomeReported', 'NetIncomeCalc')
    ]

    empty_vals = [''] * year_count
    for name, a, b in validations:
        diff_val = diff_pair(anchors.get(a), anchors.get(b))
        audit_ws.append(['Validation', name] + empty_vals + [check_flag(diff_val), _safe_float_for_excel(diff_val)])

    total_cols = 2 + len(year_cols) + 2
    for i in range(1, total_cols + 1):
        width = 28 if i == 2 else 18
        audit_ws.column_dimensions[get_column_letter(i)].width = width
    for col_idx in range(1,total_cols+1):
        audit_ws.cell(1, col_idx).font = Font(bold=True)

    try:
        image = convert_pdf_page_to_image(pdf_path, page_number=(result.get('pages_used') or [1])[0], dpi=170)
        if image is not None:
            image_path = os.path.join(REPORT_DIR, base + '_statement_page.png')
            image.save(image_path)
            img = XLImage(image_path)
            img.width = 900
            img.height = int(img.height * (900 / img.width)) if img.width else img.height
            # ============================================================
            # 📸 MOVE IMAGE TO EXTRACTED SHEET
            # ============================================================
            img.anchor = 'F2'
            ws.add_image(img)
    except Exception as e:
        meta.append(('ImageInsertError', str(e)))
    wb.save(out_path)
    return out_path

def build_batch_summary_excel(results, folder_path):
    ensure_dir(REPORT_DIR)
    out_path = os.path.join(REPORT_DIR, 'StageD_Summary.xlsx')
    df = pd.DataFrame(results)
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Summary')
        counts = df['Status'].value_counts(dropna=False).rename_axis('Status').reset_index(name='Count') if not df.empty else pd.DataFrame(columns=['Status','Count'])
        counts.to_excel(writer, index=False, sheet_name='Counts')
    return out_path




# ============================================================
# HELPER: FORMAT DIFF (K / M / B)
# ============================================================

def fmt_diff_value(diff):
    if diff is None or not math.isfinite(diff):
        return "inf"
    if diff >= 1_000_000_000:
        return f"{round(diff / 1_000_000_000, 2)}B"
    elif diff >= 1_000_000:
        return f"{round(diff / 1_000_000, 2)}M"
    elif diff >= 1_000:
        return f"{round(diff / 1_000, 2)}K"
    else:
        return str(round(diff, 2))
    
    
def detect_insurance_model(df):
    """Detect if statement is insurance-based using keywords"""

    text = " ".join(df['Item'].dropna().astype(str)).lower()

    insurance_keywords = [
        "insurance revenue",
        "insurance service expenses",
        "reinsurance",
        "insurance service result",
        "insurance finance"
    ]

    hits = sum(1 for kw in insurance_keywords if kw in text)

    print(f"🟦 Insurance Detection Hits: {hits}")

    return hits >= 2